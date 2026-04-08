from __future__ import annotations

import csv
import json
from pathlib import Path

from fastapi.encoders import jsonable_encoder
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.db.models import (
    CaseRecord,
    DeliveryDocketRecord,
    DocumentRecord,
    ExtractionRunRecord,
    ExportRecord,
    InvoiceRecord,
    ReconciliationRunRecord,
)
from app.schemas.canonical import (
    AccountingTemplateDefinition,
    CanonicalInvoice,
    DeliveryDocket,
    DocumentType,
    ReconciliationResult,
)
from app.services.export.accounting_mapper import accounting_export_mapper
from app.services.export.ocr_extract_mapper import ocr_extract_mapper
from app.services.export.pnl_template import load_builtin_pnl_template
from app.services.export.raw_ocr_mapper import raw_ocr_export_mapper
from app.services.export.raw_ocr_review_renderer import raw_ocr_review_renderer
from app.services.export.reconciliation_mapper import reconciliation_export_mapper
from app.services.storage.local import local_storage_service


class ExportService:
    def create_export(self, db: Session, *, case_id: str, export_format: str) -> ExportRecord:
        if export_format not in {"csv", "json", "reco_csv", "reco_excel", "ocr_excel", "ocr_html", "pnl_csv"}:
            raise ValueError(f"Unsupported export format '{export_format}'.")

        case = db.get(CaseRecord, case_id)
        if case is None:
            raise ValueError(f"Case '{case_id}' not found.")

        invoice_document = db.scalar(
            select(DocumentRecord)
            .where(DocumentRecord.case_id == case_id, DocumentRecord.doc_type == DocumentType.INVOICE.value)
            .order_by(DocumentRecord.created_at.desc())
        )
        docket_document = db.scalar(
            select(DocumentRecord)
            .where(DocumentRecord.case_id == case_id, DocumentRecord.doc_type == DocumentType.DELIVERY_DOCKET.value)
            .order_by(DocumentRecord.created_at.desc())
        )
        invoice_record = db.scalar(
            select(InvoiceRecord).where(InvoiceRecord.case_id == case_id).order_by(InvoiceRecord.created_at.desc())
        )
        docket_record = db.scalar(
            select(DeliveryDocketRecord)
            .where(DeliveryDocketRecord.case_id == case_id)
            .order_by(DeliveryDocketRecord.created_at.desc())
        )
        reconciliation_run = db.scalar(
            select(ReconciliationRunRecord)
            .where(ReconciliationRunRecord.case_id == case_id)
            .order_by(ReconciliationRunRecord.created_at.desc())
        )
        invoice_extraction_run = self._load_latest_extraction_run(db, document_id=invoice_document.id) if invoice_document else None
        docket_extraction_run = self._load_latest_extraction_run(db, document_id=docket_document.id) if docket_document else None

        if invoice_record is None or docket_record is None or reconciliation_run is None:
            raise ValueError("Extracted invoice, delivery docket, and reconciliation result are required.")

        invoice = CanonicalInvoice.model_validate(invoice_record.canonical_payload)
        docket = DeliveryDocket.model_validate(docket_record.canonical_payload)
        reconciliation = ReconciliationResult.model_validate(reconciliation_run.result_payload or {})

        if export_format == "reco_csv":
            rows = reconciliation_export_mapper.map_rows(invoice, docket, reconciliation)
            content_type = "text/csv"
            output_path = self._write_reconciliation_output(case_id, rows)
            export_payload = {
                "export_name": "reconciliation_export",
                "rows_preview": jsonable_encoder([row.model_dump(mode="json") for row in rows[:5]]),
            }
        elif export_format == "reco_excel":
            reco_rows = reconciliation_export_mapper.map_rows(invoice, docket, reconciliation)
            invoice_rows = ocr_extract_mapper.map_rows(invoice, docket)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            output_path = self._write_reconciliation_workbook(case_id, reco_rows, invoice_rows)
            rows = reco_rows
            export_payload = {
                "export_name": "reconciliation_workbook",
                "sheet_names": ["reco", "invoice"],
                "sheet_row_counts": {
                    "reco": len(reco_rows),
                    "invoice": len(invoice_rows),
                },
                "rows_preview": jsonable_encoder([row.model_dump(mode="json") for row in reco_rows[:5]]),
            }
        elif export_format == "ocr_excel":
            if invoice_extraction_run is None or docket_extraction_run is None:
                raise ValueError("Completed invoice and delivery docket extraction runs are required.")

            invoice_rows = raw_ocr_export_mapper.map_rows(invoice_extraction_run.provider_payload)
            docket_rows = raw_ocr_export_mapper.map_rows(docket_extraction_run.provider_payload)
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            output_path = self._write_raw_ocr_workbook(case_id, invoice_rows, docket_rows)
            rows = invoice_rows + docket_rows
            export_payload = {
                "export_name": "raw_ocr_workbook",
                "sheet_names": ["invoice", "docket"],
                "sheet_row_counts": {
                    "invoice": len(invoice_rows),
                    "docket": len(docket_rows),
                },
                "rows_preview": invoice_rows[:5],
            }
        elif export_format == "ocr_html":
            if invoice_extraction_run is None or docket_extraction_run is None:
                raise ValueError("Completed invoice and delivery docket extraction runs are required.")

            invoice_rows = raw_ocr_export_mapper.map_rows(invoice_extraction_run.provider_payload)
            docket_rows = raw_ocr_export_mapper.map_rows(docket_extraction_run.provider_payload)
            content_type = "text/html"
            output_path = self._write_raw_ocr_review_html(
                case_id,
                invoice_payload=invoice_extraction_run.provider_payload,
                docket_payload=docket_extraction_run.provider_payload,
            )
            rows = invoice_rows + docket_rows
            export_payload = {
                "export_name": "raw_ocr_review_html",
                "document_sections": ["invoice", "docket"],
                "flattened_row_counts": {
                    "invoice": len(invoice_rows),
                    "docket": len(docket_rows),
                },
                "rows_preview": invoice_rows[:5],
            }
        elif export_format == "pnl_csv":
            template = load_builtin_pnl_template()
            rows = accounting_export_mapper.map_rows(invoice, docket, reconciliation, template)
            content_type = "text/csv"
            output_path = self._write_template_csv(case_id, "pnl_export.csv", rows, template)
            export_payload = {
                "template_name": template.template_name,
                "template_notes": template.notes,
                "rows_preview": jsonable_encoder([row.model_dump(mode="json") for row in rows[:5]]),
            }
        else:
            template_document = db.scalar(
                select(DocumentRecord)
                .where(
                    DocumentRecord.case_id == case_id,
                    DocumentRecord.doc_type == DocumentType.ACCOUNTING_TEMPLATE.value,
                )
                .order_by(DocumentRecord.created_at.desc())
            )
            template = self._load_template(template_document)
            rows = accounting_export_mapper.map_rows(invoice, docket, reconciliation, template)
            content_type = "text/csv" if export_format == "csv" else "application/json"
            output_path = self._write_output(case_id, export_format, rows, template)
            export_payload = {
                "template_name": template.template_name,
                "template_notes": template.notes,
                "rows_preview": jsonable_encoder([row.model_dump(mode="json") for row in rows[:5]]),
            }

        export_record = ExportRecord(
            case_id=case_id,
            reconciliation_run_id=reconciliation_run.id,
            export_format=export_format,
            status="created",
            content_type=content_type,
            output_path=output_path.relative_to(local_storage_service.root).as_posix(),
            row_count=len(rows),
            export_payload=export_payload,
        )
        db.add(export_record)
        case.status = "exported" if reconciliation.approved else case.status
        db.commit()
        db.refresh(export_record)
        return export_record

    def _load_template(self, document: DocumentRecord | None) -> AccountingTemplateDefinition:
        if document and document.latest_extraction_payload:
            return AccountingTemplateDefinition.model_validate(document.latest_extraction_payload)
        return load_builtin_pnl_template()

    def _write_output(
        self,
        case_id: str,
        export_format: str,
        rows,
        template: AccountingTemplateDefinition,
    ) -> Path:
        if export_format == "csv":
            return self._write_template_csv(case_id, "accounting_export.csv", rows, template)

        output_path = local_storage_service.build_export_path(case_id, "accounting_export.json")
        payload = {
            "template_name": template.template_name,
            "template_notes": template.notes,
            "rows": [row.model_dump(mode="json") for row in rows],
        }
        output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        return output_path

    def _write_reconciliation_output(self, case_id: str, rows) -> Path:
        output_path = local_storage_service.build_export_path(case_id, "reconciliation_export.csv")
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=reconciliation_export_mapper.columns)
            writer.writeheader()
            for row in rows:
                writer.writerow(reconciliation_export_mapper.to_csv_row(row))
        return output_path

    def _write_reconciliation_workbook(
        self,
        case_id: str,
        reco_rows,
        invoice_rows: list[dict[str, str]],
    ) -> Path:
        output_path = local_storage_service.build_export_path(case_id, "reconciliation_export.xlsx")
        workbook = Workbook()
        reco_sheet = workbook.active
        reco_sheet.title = "reco"
        self._write_sheet(
            reco_sheet,
            reconciliation_export_mapper.columns,
            [reconciliation_export_mapper.to_csv_row(row) for row in reco_rows],
        )

        invoice_sheet = workbook.create_sheet(title="invoice")
        self._write_sheet(invoice_sheet, ocr_extract_mapper.columns, invoice_rows)
        workbook.save(output_path)
        return output_path

    def _write_raw_ocr_workbook(
        self,
        case_id: str,
        invoice_rows: list[dict[str, str]],
        docket_rows: list[dict[str, str]],
    ) -> Path:
        output_path = local_storage_service.build_export_path(case_id, "raw_ocr_export.xlsx")
        workbook = Workbook()
        invoice_sheet = workbook.active
        invoice_sheet.title = "invoice"
        self._write_sheet(invoice_sheet, raw_ocr_export_mapper.columns, invoice_rows)

        docket_sheet = workbook.create_sheet(title="docket")
        self._write_sheet(docket_sheet, raw_ocr_export_mapper.columns, docket_rows)
        workbook.save(output_path)
        return output_path

    def _write_raw_ocr_review_html(
        self,
        case_id: str,
        *,
        invoice_payload: dict[str, object] | None,
        docket_payload: dict[str, object] | None,
    ) -> Path:
        output_path = local_storage_service.build_export_path(case_id, "raw_ocr_review.html")
        html = raw_ocr_review_renderer.render(
            case_id=case_id,
            invoice_payload=invoice_payload,
            docket_payload=docket_payload,
        )
        output_path.write_text(html, encoding="utf-8")
        return output_path

    def _write_template_csv(
        self,
        case_id: str,
        filename: str,
        rows,
        template: AccountingTemplateDefinition,
    ) -> Path:
        output_path = local_storage_service.build_export_path(case_id, filename)
        columns = [column.column_name for column in template.columns]
        if not columns and rows:
            columns = list(rows[0].template_values)
        with output_path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns)
            writer.writeheader()
            for row in rows:
                writer.writerow({key: row.template_values.get(key, "") for key in columns})
        return output_path

    def _write_sheet(self, sheet, columns: list[str], rows: list[dict[str, str]]) -> None:
        sheet.append(columns)
        for cell in sheet[1]:
            cell.font = Font(bold=True)

        for row in rows:
            sheet.append([row.get(column, "") for column in columns])

        sheet.freeze_panes = "A2"

    def _load_latest_extraction_run(self, db: Session, *, document_id: str) -> ExtractionRunRecord | None:
        return db.scalar(
            select(ExtractionRunRecord)
            .where(ExtractionRunRecord.document_id == document_id, ExtractionRunRecord.status == "completed")
            .order_by(ExtractionRunRecord.created_at.desc())
        )


export_service = ExportService()
