from __future__ import annotations

from io import BytesIO

from openpyxl import load_workbook


def test_full_api_flow(client, upload_payloads):
    response = client.post("/api/cases/uploads", files=upload_payloads)
    assert response.status_code == 200
    case = response.json()
    case_id = case["id"]

    extract = client.post(
        f"/api/cases/{case_id}/extract",
        json={"provider_name": "mock", "force": True},
    )
    assert extract.status_code == 200
    assert len(extract.json()["runs"]) == 2

    reconcile = client.post(f"/api/cases/{case_id}/reconcile", json={})
    assert reconcile.status_code == 200
    assert reconcile.json()["status"] == "exceptions"

    reco_export = client.post(f"/api/exports/cases/{case_id}", json={"export_format": "reco_excel"})
    assert reco_export.status_code == 200
    reco_export_payload = reco_export.json()
    assert reco_export_payload["row_count"] == 8
    assert reco_export_payload["export_payload"]["sheet_names"] == ["reco", "invoice"]

    reco_download = client.get(f"/api/exports/{reco_export_payload['id']}/download")
    assert reco_download.status_code == 200
    workbook = load_workbook(BytesIO(reco_download.content))
    assert workbook.sheetnames == ["reco", "invoice"]
    reco_sheet = workbook["reco"]
    invoice_sheet = workbook["invoice"]
    assert reco_sheet["A1"].value == "Product Code"
    assert reco_sheet["H4"].value == "Quantity mismatch; Amount mismatch"
    assert invoice_sheet["A1"].value == "Source Document"
    assert invoice_sheet["A2"].value == "invoice"
    assert any(cell.value == "delivery_docket" for cell in invoice_sheet["A"])

    raw_ocr_export = client.post(f"/api/exports/cases/{case_id}", json={"export_format": "ocr_excel"})
    assert raw_ocr_export.status_code == 200
    raw_ocr_export_payload = raw_ocr_export.json()
    assert raw_ocr_export_payload["export_payload"]["sheet_names"] == ["invoice", "docket"]
    assert raw_ocr_export_payload["export_payload"]["sheet_row_counts"]["invoice"] > 0
    assert raw_ocr_export_payload["export_payload"]["sheet_row_counts"]["docket"] > 0

    raw_ocr_download = client.get(f"/api/exports/{raw_ocr_export_payload['id']}/download")
    assert raw_ocr_download.status_code == 200
    raw_workbook = load_workbook(BytesIO(raw_ocr_download.content))
    assert raw_workbook.sheetnames == ["invoice", "docket"]
    raw_invoice_sheet = raw_workbook["invoice"]
    raw_docket_sheet = raw_workbook["docket"]
    assert raw_invoice_sheet["A1"].value == "JSON Path"
    assert raw_docket_sheet["A1"].value == "JSON Path"

    invoice_values = {
        row[0]: row[1]
        for row in raw_invoice_sheet.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    docket_values = {
        row[0]: row[1]
        for row in raw_docket_sheet.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    assert invoice_values["source_filename"] == "Invoice_598527_Account_64876_Division_MRPI_Full_unlocked.pdf"
    assert docket_values["source_filename"] == "Delivery Docket.jpeg"
    assert "Mock OCR fixture" in invoice_values["notes[0]"]
    assert "Mock OCR fixture" in docket_values["notes[0]"]

    raw_ocr_html_export = client.post(f"/api/exports/cases/{case_id}", json={"export_format": "ocr_html"})
    assert raw_ocr_html_export.status_code == 200
    raw_ocr_html_payload = raw_ocr_html_export.json()
    assert raw_ocr_html_payload["export_payload"]["document_sections"] == ["invoice", "docket"]
    assert raw_ocr_html_payload["export_payload"]["flattened_row_counts"]["invoice"] > 0
    assert raw_ocr_html_payload["export_payload"]["flattened_row_counts"]["docket"] > 0

    raw_ocr_html_download = client.get(f"/api/exports/{raw_ocr_html_payload['id']}/download")
    assert raw_ocr_html_download.status_code == 200
    assert "text/html" in raw_ocr_html_download.headers["content-type"]
    assert "Invoice OCR Review" in raw_ocr_html_download.text
    assert "Delivery Docket OCR Review" in raw_ocr_html_download.text
    assert "Flattened raw payload audit" in raw_ocr_html_download.text
    assert "Invoice_598527_Account_64876_Division_MRPI_Full_unlocked.pdf" in raw_ocr_html_download.text

    pnl_export = client.post(f"/api/exports/cases/{case_id}", json={"export_format": "pnl_csv"})
    assert pnl_export.status_code == 200
    pnl_export_payload = pnl_export.json()
    assert pnl_export_payload["row_count"] == 8
    assert pnl_export_payload["export_payload"]["template_name"] == "Built-in P&L Purchase Template"

    pnl_download = client.get(f"/api/exports/{pnl_export_payload['id']}/download")
    assert pnl_download.status_code == 200
    assert "P&L Section" in pnl_download.text
    assert "P&L Notes" in pnl_download.text
    assert "Final Comment" in pnl_download.text

    case_detail = client.get(f"/api/cases/{case_id}")
    assert case_detail.status_code == 200
    detail = case_detail.json()
    assert detail["document_count"] == 2
    assert detail["invoice"]["header"]["invoice_number"] == "598527"
    assert detail["delivery_docket"]["docket_number"] == "DD-240326-2064"


def test_invoice_and_docket_edits_persist(client, upload_payloads):
    response = client.post("/api/cases/uploads", files=upload_payloads)
    assert response.status_code == 200
    case_id = response.json()["id"]

    extract = client.post(
        f"/api/cases/{case_id}/extract",
        json={"provider_name": "mock", "force": True},
    )
    assert extract.status_code == 200

    invoice_update = client.put(
        f"/api/cases/{case_id}/invoice",
        json={
            "rows": [
                {
                    "supplier": "Edited Supplier",
                    "product_code": "SKU-001",
                    "product_name": "Edited Apples",
                    "quantity_invoice": "12",
                    "pre_amount_invoice": "24.50",
                    "vat_invoice": "5.64",
                    "total_invoice": "30.14",
                }
            ]
        },
    )
    assert invoice_update.status_code == 200
    invoice_payload = invoice_update.json()["payload"]
    assert invoice_payload["supplier"]["name"] == "Edited Supplier"
    assert invoice_payload["header"]["subtotal_amount"] == "24.50"
    assert invoice_payload["lines"][0]["product_code"] == "SKU-001"
    assert invoice_payload["lines"][0]["gross_amount"] == "30.14"

    docket_update = client.put(
        f"/api/cases/{case_id}/delivery-docket",
        json={
            "rows": [
                {
                    "supplier": "Edited Supplier",
                    "product_code": "SKU-001",
                    "product_name": "Edited Apples",
                    "quantity_docket": "10",
                    "amount_docket": "",
                }
            ]
        },
    )
    assert docket_update.status_code == 200
    docket_payload = docket_update.json()["payload"]
    assert docket_payload["supplier_name"] == "Edited Supplier"
    assert docket_payload["lines"][0]["product_code"] == "SKU-001"
    assert docket_payload["lines"][0]["quantity_delivered"] == "10"
    assert docket_payload["lines"][0]["extended_amount"] is None


def test_manual_reconciliation_persists_latest_run_and_export_order(client, upload_payloads):
    response = client.post("/api/cases/uploads", files=upload_payloads)
    assert response.status_code == 200
    case_id = response.json()["id"]

    extract = client.post(
        f"/api/cases/{case_id}/extract",
        json={"provider_name": "mock", "force": True},
    )
    assert extract.status_code == 200

    invoice_payload = client.get(f"/api/cases/{case_id}/invoice").json()["payload"]
    line_one_name = invoice_payload["lines"][0]["description"]
    line_two_name = invoice_payload["lines"][1]["description"]

    auto_reconcile = client.post(f"/api/cases/{case_id}/reconcile", json={})
    assert auto_reconcile.status_code == 200
    auto_run_id = auto_reconcile.json()["id"]

    manual_reconcile = client.post(
        f"/api/cases/{case_id}/reconciliation/manual",
        json={
            "base_reconciliation_run_id": auto_run_id,
            "pairs": [
                {"invoice_line_number": 2, "docket_line_number": 2, "position": 0},
                {"invoice_line_number": 1, "docket_line_number": 1, "position": 1},
            ],
        },
    )
    assert manual_reconcile.status_code == 200
    manual_payload = manual_reconcile.json()
    merged_lines = [
        line
        for line in manual_payload["result_payload"]["reconciled_lines"]
        if line["invoice_line_number"] is not None and line["docket_line_number"] is not None
    ]
    assert merged_lines[0]["invoice_line_number"] == 2
    assert merged_lines[0]["manual_pair_position"] == 0
    assert merged_lines[0]["match_origin"] == "auto"

    reco_export = client.post(f"/api/exports/cases/{case_id}", json={"export_format": "reco_excel"})
    assert reco_export.status_code == 200
    reco_download = client.get(f"/api/exports/{reco_export.json()['id']}/download")
    workbook = load_workbook(BytesIO(reco_download.content))
    reco_sheet = workbook["reco"]
    assert reco_sheet["B2"].value == line_two_name
    assert reco_sheet["B3"].value == line_one_name


def test_manual_reconciliation_rejects_stale_and_duplicate_pairs(client, upload_payloads):
    response = client.post("/api/cases/uploads", files=upload_payloads)
    assert response.status_code == 200
    case_id = response.json()["id"]

    extract = client.post(
        f"/api/cases/{case_id}/extract",
        json={"provider_name": "mock", "force": True},
    )
    assert extract.status_code == 200

    auto_reconcile = client.post(f"/api/cases/{case_id}/reconcile", json={})
    assert auto_reconcile.status_code == 200
    auto_run_id = auto_reconcile.json()["id"]

    stale_manual = client.post(
        f"/api/cases/{case_id}/reconciliation/manual",
        json={"base_reconciliation_run_id": "stale-run-id", "pairs": []},
    )
    assert stale_manual.status_code == 400
    assert "stale" in stale_manual.json()["detail"].lower()

    duplicate_manual = client.post(
        f"/api/cases/{case_id}/reconciliation/manual",
        json={
            "base_reconciliation_run_id": auto_run_id,
            "pairs": [
                {"invoice_line_number": 1, "docket_line_number": 1, "position": 0},
                {"invoice_line_number": 1, "docket_line_number": 2, "position": 1},
            ],
        },
    )
    assert duplicate_manual.status_code == 400
    assert "cannot be paired more than once" in duplicate_manual.json()["detail"].lower()
